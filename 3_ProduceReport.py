# Create a summary report of produce and total sold as shown below:


import openpyxl as xl

wb = xl.load_workbook('produceSales.xlsx')


ws = wb.active

maxC = ws.max_column
maxR = ws.max_row

produce_dict = {}

##print("Produce","\t","Cost Per Pound","\t","Amt Sold","\t","Total")
    
for currentrow in ws.iter_rows(min_row = 2, max_row=maxR , max_col=maxC):
    name = currentrow[0].value
    cost = float(currentrow[1].value)
    amt_sold = round(float(currentrow[2].value),2)
    total = round(float(currentrow[3].value),2)

    produce_dict.setdefault(name,[0.00,0.00,0.00])

    produce_dict[name][0] = round(cost,2)

    
    produce_dict[name][1] += round(amt_sold,2)

    produce_dict[name][2] += round(total,2)



    #print(produce_dict)
    #input()

for k,v in produce_dict.items():
    produce_dict[k] = [v[0],round(v[1]),round(v[2])]

#output to the screen

import pandas as pd

produce = pd.DataFrame(produce_dict)

produce.index = ['cost', 'quantity', 'total sales']

#produce that had the highest and lowest in total sales
produce.sort_values(by='total sales',axis=1, ascending=True)
lowest = produce.loc[['total sales'],['Watermelon']]
print("lowest", lowest)
highest = produce.loc[['total sales'],['Cherries']]
print("highest", highest)

#display quantity and total sales for Oranges and Beets
print(produce.loc[['quantity','total sales'], ['Orange','Beets']])

#display total sales for cucumbers
print("cucumber total sales:", produce.at['total sales','Cucumber'])

#create a new dataframe for only produce sold between 11500-12000
produce = produce.T
produce = produce.sort_values(by=['quantity'], axis=0, ascending=False)
produce2 = pd.DataFrame()
produce2 = produce.loc[(produce['quantity'] <= 12000) & (produce['quantity'] >= 11500)]
print("new dataframe")
print(produce2)

#transpose
print("transposed new dataframe")
print(produce2.T)
